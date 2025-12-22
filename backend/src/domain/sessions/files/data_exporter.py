"""
Data Exporter

Specialized generator for data export files (CSV, JSON, Excel, XML).
"""

from typing import Optional, Dict, Any, List, Union
from pathlib import Path
from datetime import datetime
import uuid
import json
import csv
import io
import logging

from ..models import Attachment, AttachmentType
from .generator import BaseGenerator
from .types import (
    GenerationType,
    GenerationRequest,
    GenerationResult,
    ExportFormat,
    CONTENT_TYPE_MAP
)

logger = logging.getLogger(__name__)


class DataExporter(BaseGenerator):
    """資料匯出器

    匯出各種格式的資料文件。

    功能:
    - CSV 匯出
    - JSON 匯出
    - Excel 匯出
    - XML 匯出
    - 資料轉換
    """

    # 格式對應的擴展名
    FORMAT_TO_EXT = {
        ExportFormat.CSV: ".csv",
        ExportFormat.JSON: ".json",
        ExportFormat.EXCEL: ".xlsx",
        ExportFormat.XML: ".xml"
    }

    # 格式對應的 MIME 類型
    FORMAT_TO_CONTENT_TYPE = {
        ExportFormat.CSV: "text/csv",
        ExportFormat.JSON: "application/json",
        ExportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ExportFormat.XML: "application/xml"
    }

    def __init__(self, storage: Optional[Any] = None):
        """
        初始化資料匯出器

        Args:
            storage: 存儲服務實例
        """
        self._storage = storage

    def supports(self, generation_type: GenerationType) -> bool:
        """檢查是否支援此生成類型"""
        return generation_type == GenerationType.DATA

    async def generate(
        self,
        session_id: str,
        request: GenerationRequest
    ) -> GenerationResult:
        """
        生成資料匯出文件

        Args:
            session_id: Session ID
            request: 生成請求

        Returns:
            GenerationResult: 生成結果
        """
        try:
            # 確定輸出格式
            export_format = request.export_format or ExportFormat.CSV
            extension = self.FORMAT_TO_EXT.get(export_format, ".csv")
            content_type = self.FORMAT_TO_CONTENT_TYPE.get(export_format, "text/csv")

            # 生成文件名
            filename = request.filename
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"export_{timestamp}{extension}"
            elif not filename.endswith(extension):
                filename = f"{filename}{extension}"

            # 解析資料
            data = self._parse_data(request.content, request.options)

            # 根據格式生成內容
            if export_format == ExportFormat.CSV:
                content = self._export_csv(data, request.options)
            elif export_format == ExportFormat.JSON:
                content = self._export_json(data, request.options)
            elif export_format == ExportFormat.EXCEL:
                content = await self._export_excel(data, request.options)
            elif export_format == ExportFormat.XML:
                content = self._export_xml(data, request.options)
            else:
                content = self._export_csv(data, request.options)

            # 創建附件
            attachment_id = str(uuid.uuid4())

            # 計算大小
            if isinstance(content, bytes):
                size = len(content)
            else:
                size = len(content.encode("utf-8"))

            # 存儲文件
            if self._storage:
                storage_path = await self._storage.store_content(
                    session_id=session_id,
                    attachment_id=attachment_id,
                    content=content,
                    filename=filename
                )
            else:
                storage_path = f"/tmp/sessions/{session_id}/{attachment_id}/{filename}"

            logger.info(f"Exported data: {filename} ({size} bytes)")

            return GenerationResult(
                success=True,
                generation_type=GenerationType.DATA,
                attachment_id=attachment_id,
                filename=filename,
                content_type=content_type,
                size=size
            )

        except Exception as e:
            logger.error(f"Data export failed: {e}")
            return GenerationResult(
                success=False,
                generation_type=GenerationType.DATA,
                error=str(e)
            )

    def _parse_data(
        self,
        content: str,
        options: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """解析輸入資料"""
        # 如果已經是列表，直接使用
        if isinstance(content, list):
            return content

        # 如果是字典，包裝成列表
        if isinstance(content, dict):
            return [content]

        # 嘗試解析 JSON
        try:
            data = json.loads(content)
            if isinstance(data, list):
                return data
            elif isinstance(data, dict):
                # 如果是包含 data 鍵的字典
                if "data" in data and isinstance(data["data"], list):
                    return data["data"]
                return [data]
        except json.JSONDecodeError:
            pass

        # 嘗試解析 CSV 格式字符串
        try:
            reader = csv.DictReader(io.StringIO(content))
            return list(reader)
        except Exception:
            pass

        # 無法解析，返回空列表
        logger.warning("Could not parse data content, returning empty list")
        return []

    def _export_csv(
        self,
        data: List[Dict[str, Any]],
        options: Dict[str, Any]
    ) -> str:
        """匯出為 CSV"""
        if not data:
            return ""

        # 獲取所有欄位
        fieldnames = options.get("columns")
        if not fieldnames:
            # 從所有記錄中收集欄位
            fieldnames = []
            for row in data:
                for key in row.keys():
                    if key not in fieldnames:
                        fieldnames.append(key)

        # 生成 CSV
        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=fieldnames,
            extrasaction='ignore',
            quoting=csv.QUOTE_MINIMAL
        )

        # 寫入標題行 (如果選項中啟用)
        if options.get("include_header", True):
            writer.writeheader()

        # 寫入資料行
        for row in data:
            writer.writerow(row)

        return output.getvalue()

    def _export_json(
        self,
        data: List[Dict[str, Any]],
        options: Dict[str, Any]
    ) -> str:
        """匯出為 JSON"""
        indent = options.get("indent", 2) if options.get("pretty", True) else None
        ensure_ascii = options.get("ensure_ascii", False)

        # 包裝格式
        if options.get("wrap_data", False):
            output = {
                "data": data,
                "count": len(data),
                "exported_at": datetime.now().isoformat()
            }
        else:
            output = data

        return json.dumps(output, indent=indent, ensure_ascii=ensure_ascii, default=str)

    async def _export_excel(
        self,
        data: List[Dict[str, Any]],
        options: Dict[str, Any]
    ) -> bytes:
        """匯出為 Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return self._export_csv(data, options).encode("utf-8")

        if not data:
            # 創建空白 Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        # 獲取欄位
        fieldnames = options.get("columns")
        if not fieldnames:
            fieldnames = []
            for row in data:
                for key in row.keys():
                    if key not in fieldnames:
                        fieldnames.append(key)

        # 創建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = options.get("sheet_name", "Data")

        # 定義樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 寫入標題行
        for col, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 寫入資料行
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                value = row_data.get(field, "")
                # 處理複雜類型
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 自動調整列寬
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # 凍結首行
        if options.get("freeze_header", True):
            ws.freeze_panes = "A2"

        # 保存到 bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _export_xml(
        self,
        data: List[Dict[str, Any]],
        options: Dict[str, Any]
    ) -> str:
        """匯出為 XML"""
        root_element = options.get("root_element", "data")
        item_element = options.get("item_element", "item")
        indent = "  " if options.get("pretty", True) else ""
        newline = "\n" if options.get("pretty", True) else ""

        lines = [
            '<?xml version="1.0" encoding="UTF-8"?>',
            f'<{root_element}>'
        ]

        for item in data:
            lines.append(f'{indent}<{item_element}>')
            for key, value in item.items():
                # 處理複雜類型
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                # XML 轉義
                value = self._xml_escape(str(value) if value is not None else "")
                key = self._xml_safe_tag(key)
                lines.append(f'{indent}{indent}<{key}>{value}</{key}>')
            lines.append(f'{indent}</{item_element}>')

        lines.append(f'</{root_element}>')

        return newline.join(lines)

    def _xml_escape(self, text: str) -> str:
        """XML 字符轉義"""
        return (text
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&apos;"))

    def _xml_safe_tag(self, tag: str) -> str:
        """確保 XML 標籤名稱有效"""
        import re
        # 移除無效字符，只保留字母、數字、下劃線
        safe_tag = re.sub(r'[^a-zA-Z0-9_]', '_', tag)
        # 確保不以數字開頭
        if safe_tag and safe_tag[0].isdigit():
            safe_tag = f"_{safe_tag}"
        return safe_tag or "field"

    async def export_data(
        self,
        session_id: str,
        data: Union[List[Dict[str, Any]], str],
        filename: str,
        format: ExportFormat = ExportFormat.CSV,
        columns: Optional[List[str]] = None,
        include_header: bool = True,
        pretty: bool = True
    ) -> GenerationResult:
        """
        便捷方法：匯出資料

        Args:
            session_id: Session ID
            data: 資料內容 (列表或 JSON 字符串)
            filename: 文件名
            format: 輸出格式
            columns: 指定匯出欄位
            include_header: 是否包含標題行
            pretty: 是否美化輸出

        Returns:
            GenerationResult: 生成結果
        """
        content = data if isinstance(data, str) else json.dumps(data)

        request = GenerationRequest(
            generation_type=GenerationType.DATA,
            content=content,
            filename=filename,
            export_format=format,
            options={
                "columns": columns,
                "include_header": include_header,
                "pretty": pretty
            }
        )

        return await self.generate(session_id, request)

    async def export_analysis_results(
        self,
        session_id: str,
        analysis_data: Dict[str, Any],
        filename: str = "analysis_results",
        format: ExportFormat = ExportFormat.JSON
    ) -> GenerationResult:
        """
        匯出分析結果

        Args:
            session_id: Session ID
            analysis_data: 分析結果資料
            filename: 文件名
            format: 輸出格式

        Returns:
            GenerationResult: 生成結果
        """
        # 結構化分析結果
        export_data = {
            "analysis_id": str(uuid.uuid4()),
            "session_id": session_id,
            "exported_at": datetime.now().isoformat(),
            "results": analysis_data
        }

        return await self.export_data(
            session_id=session_id,
            data=json.dumps(export_data, default=str),
            filename=filename,
            format=format,
            pretty=True
        )

    async def export_table(
        self,
        session_id: str,
        headers: List[str],
        rows: List[List[Any]],
        filename: str = "table_data",
        format: ExportFormat = ExportFormat.CSV
    ) -> GenerationResult:
        """
        匯出表格資料

        Args:
            session_id: Session ID
            headers: 表頭列表
            rows: 資料行列表
            filename: 文件名
            format: 輸出格式

        Returns:
            GenerationResult: 生成結果
        """
        # 將表格轉換為字典列表
        data = []
        for row in rows:
            row_dict = {}
            for i, header in enumerate(headers):
                row_dict[header] = row[i] if i < len(row) else ""
            data.append(row_dict)

        return await self.export_data(
            session_id=session_id,
            data=data,
            filename=filename,
            format=format,
            columns=headers
        )

    @classmethod
    def get_supported_formats(cls) -> List[ExportFormat]:
        """獲取支援的匯出格式"""
        return list(cls.FORMAT_TO_EXT.keys())
